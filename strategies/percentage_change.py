import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def use_percentage_change_strategy(df,stock_symbol):
    # Calculate percentage change between consecutive days
    df['pct_change'] = df['close'].pct_change() * 100
    
    # Generate trading signals
    df['position'] = 0
    df['position'][df['pct_change'] < 0] = 1   # Buy signal
    df['position'][df['pct_change'] > 0] = -1   # Sell signal
    
    # Ensure the 'date' column is in datetime format
    df['date'] = pd.to_datetime(df['date'])
    # Initial capital and variables to keep track of cash and shares
    initial_capital = 10000
    shares = 0
    cash = initial_capital
    trade_log = []
    df['portfolio_value'] = initial_capital
    # Iterate through the DataFrame to simulate trading
    for i in range(1, len(df)):
        if df['position'].iloc[i] == 1:
            # Buy signal: Calculate how many shares to buy and update cash
            amount_to_invest = (abs(df['pct_change'].iloc[i]) / 100) * cash
            shares_to_buy = amount_to_invest // df['close'].iloc[i]
            cash -= shares_to_buy * df['close'].iloc[i]
            shares += shares_to_buy
            trade_log.append({
                'Date': df['date'].iloc[i],
                'Type': 'Buy',
                'Shares': shares_to_buy,
                'Price': df['close'].iloc[i],
                'Cash': cash
            })
        elif df['position'].iloc[i] == -1:
            # Sell signal: Sell all shares and update cash
            shares_to_sell = int((abs(df['pct_change'].iloc[i]) / 100) * shares)
            cash += shares_to_sell * df['close'].iloc[i]
            shares -= shares_to_sell
            trade_log.append({
                'Date': df['date'].iloc[i],
                'Type': 'Sell',
                'Shares': shares_to_sell,
                'Price': df['close'].iloc[i],
                'Cash': cash
            })
        df['portfolio_value'].iloc[i] = cash + shares * df['close'].iloc[i]
            

    # Calculate final portfolio value
    final_value = cash + shares * df['close'].iloc[-1]
    profit_loss = final_value - initial_capital
    trade_log_df = pd.DataFrame(trade_log)

    # Generate an Excel report
    generate_excel_report(df, trade_log_df, final_value, profit_loss, initial_capital, cash, shares,stock_symbol)

    return final_value, profit_loss

def generate_excel_report(df, trade_log_df, final_value, profit_loss, initial_capital, cash, shares,stock_symbol):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"

    # Summary Sheet
    ws1["A1"] = "Initial Capital"
    ws1["B1"] = initial_capital
    ws1["A2"] = "Final Portfolio Value"
    ws1["B2"] = final_value
    ws1["A3"] = "Total Profit/Loss"
    ws1["B3"] = profit_loss

    # Trade Log Sheet
    ws2 = wb.create_sheet(title="Trade Log")
    for r in dataframe_to_rows(trade_log_df, index=False, header=True):
        ws2.append(r)
    
    for cell in ws2["1:1"]:
        cell.font = Font(bold=True)

    # Format the Date column in the Trade Log sheet
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'  # Set the date format

    # Equity Curve Sheet
    ws3 = wb.create_sheet(title="Equity Curve")
    for r in dataframe_to_rows(df[['date', 'portfolio_value']], index=False, header=True):
        ws3.append(r)
    
    for cell in ws3["1:1"]:
        cell.font = Font(bold=True)

    # Format the Date column in the Equity Curve sheet
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'  # Set the date format

    # Save Excel file
    report_filename = f"reports/backtest_report_{stock_symbol}_percentage_change_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(report_filename)
    print(f"Report saved to {report_filename}")